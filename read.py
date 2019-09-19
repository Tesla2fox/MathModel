
import  openpyxl
import sympy as sp
import enum
from  GreedPlan import  Greed_Plan,CorType
# import  GreedPlan

import plotly.graph_objects as go
import numpy as np
from  enum import Enum
import plotly
import readcfg as rd
import math


if __name__ == '__main__':
    case1 = True
    if  case1 :
        wb = openpyxl.load_workbook('.//data//附件1：数据集1-终稿.xlsx')
        print('ins 1')
        '''
        ins = 1
        '''
        a1 = 25
        a2 = 20
        b1 = 20
        b2 = 25
        th = 30
    else:
        wb = openpyxl.load_workbook('.//data//附件2：数据集2-终稿.xlsx')
        print('ins 2')
        '''
        ins = 2
        '''
        a1 = 20
        a2 = 10
        b1 = 15
        b2 = 20
        th = 20

    ws = wb.active
    MAX_ROW = ws.max_row
    MAX_COL = ws.max_column

    print('max_row =', MAX_ROW)
    print('max_col =', MAX_COL)

    # print(ws)
    # print(ws['A1'])
    # print('xx')

    pnt_A = sp.Point3D(ws['B3'].value, ws['C3'].value, ws['D3'].value)
    pnt_B = sp.Point3D(ws['B'+ str(MAX_ROW)].value, ws['C' + str(MAX_ROW)].value, ws['D'+ str(MAX_ROW)].value)



    lst_x = []
    lst_y = []
    lst_z = []

    pntLst = []
    typeLst = []
    typeValue = []
    pTypeLst = []
    pntLst.append(pnt_A)
    for i in range(4,MAX_ROW):
        lst_x.append(ws['B'+str(i)].value)
        lst_y.append(ws['C'+str(i)].value)
        lst_z.append(ws['D'+str(i)].value)
        pntLst.append(sp.Point3D(lst_x[-1],lst_y[-1],lst_z[-1]))
        if ws['E'+ str(i)].value == 1:
            typeLst.append(CorType.Vertical)
            typeValue.append(1)
        else:
            typeLst.append(CorType.Horizontal)
            typeValue.append(0)
        pTypeLst.append(ws['F'+str(i)].value)

    pntLst.append(pnt_B)

    f_data = open('.//data//ins_1_test.dat' , 'w')
    rd.writeConf(f_data,'x',lst_x)
    rd.writeConf(f_data,'y',lst_y)
    rd.writeConf(f_data,'z',lst_z)
    rd.writeConf(f_data,'type',typeValue)
    rd.writeConf(f_data, 'ptype', pTypeLst)
    rd.writeConf(f_data,'pnt_A', [ws['B3'].value, ws['C3'].value, ws['D3'].value])
    rd.writeConf(f_data,'pnt_B', [ws['B'+ str(MAX_ROW)].value, ws['C' + str(MAX_ROW)].value, ws['D'+ str(MAX_ROW)].value])
    # f_data.write('pnt_a_x ',ws['B3'].value)

    # self._disMat = np.zeros((len(self._pntLst), len(self._pntLst)))
    print(len(pntLst))

    dis_v_lst = []
    dis_h_lst = []
    dis_lst = []
    for i in range(len(pntLst)):
        for j in range(len(pntLst)):
            if i == j:
                dis_v = 0
                dis_h = 0
                dis = 0
                # self._disMat[i][j] = 0
            else:
                dis_v = abs(pntLst[i].z - pntLst[j].z)
                dis_h = math.sqrt((pntLst[i].x - pntLst[j].x)**2  +  (pntLst[i].y - pntLst[j].y)**2)
                dis = math.sqrt((pntLst[i].x - pntLst[j].x)**2  +  (pntLst[i].y - pntLst[j].y)**2 + (pntLst[i].z - pntLst[j].z)**2)
            dis_v_lst.append(dis_v)
            dis_h_lst.append(dis_h)
            dis_lst.append(dis)

    rd.writeConf(f_data, 'dis_v', dis_v_lst)
    rd.writeConf(f_data, 'dis_h', dis_h_lst)
    rd.writeConf(f_data, 'dis', dis_lst)

            # self._pntLst[i].distance(self._pntLst[j])
            # self._disMat[i][j] = dis
    # return True


    f_data.close()

    # print(typeLst)
    exit()






    gp = Greed_Plan(pntLst , typeLst, pnt_A, pnt_B , a1 ,a2 ,b1 ,b2 ,th)
    gp.plan()
    # print(lst_x)
    # print(len(lst_x))